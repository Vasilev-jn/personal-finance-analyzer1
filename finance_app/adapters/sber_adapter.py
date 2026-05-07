# -*- coding: utf-8 -*-
from datetime import date
from decimal import Decimal
from typing import Dict, List, Optional
from uuid import uuid4

from openpyxl import load_workbook

from finance_app.domain import Operation, OperationType, Vault
from finance_app.utils import normalize_text, parse_decimal


MONTHS = {
    "янв": 1,
    "января": 1,
    "фев": 2,
    "февраля": 2,
    "мар": 3,
    "марта": 3,
    "апр": 4,
    "апреля": 4,
    "май": 5,
    "мая": 5,
    "июн": 6,
    "июня": 6,
    "июл": 7,
    "июля": 7,
    "авг": 8,
    "августа": 8,
    "сен": 9,
    "сент": 9,
    "сентября": 9,
    "окт": 10,
    "октября": 10,
    "нояб": 11,
    "ноября": 11,
    "дек": 12,
    "декабря": 12,
}


def _push_skip(report: Dict[str, object], row_index: int, reason: str) -> None:
    report["skipped"] = int(report.get("skipped", 0)) + 1
    errors = report.setdefault("errors", [])
    if isinstance(errors, list) and len(errors) < 30:
        errors.append({"row": row_index, "reason": reason})


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_amount(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    return parse_decimal(str(value))


def _parse_sber_date(value: object) -> Optional[date]:
    if isinstance(value, date):
        return value
    raw = _text(value).replace(",", "")
    parts = raw.split()
    if len(parts) < 3:
        return None
    try:
        day = int(parts[0])
        month_key = parts[1].lower().replace(".", "")
        month = MONTHS.get(month_key)
        year = int(parts[2])
        if not month:
            return None
        return date(year, month, day)
    except Exception:
        return None


def _detect_type(type_raw: str, amount: Decimal) -> OperationType:
    value = normalize_text(type_raw)
    if any(key in value for key in ("доход", "поступ", "пополн", "зачисл", "возврат")):
        return OperationType.INCOME
    if any(key in value for key in ("расход", "списан", "покупк", "оплат")):
        return OperationType.EXPENSE
    return OperationType.INCOME if amount >= 0 else OperationType.EXPENSE


def _header_index(row: tuple[object, ...]) -> Dict[str, int]:
    return {_text(value): idx for idx, value in enumerate(row) if _text(value)}


def _cell(row: tuple[object, ...], headers: Dict[str, int], name: str) -> object:
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_inactive(state_raw: str) -> bool:
    state = normalize_text(state_raw)
    return any(key in state for key in ("отмен", "отклон", "неуспеш"))


def import_sber_xlsx(
    vault: Vault,
    path: str,
    file_id: str,
    report: Optional[Dict[str, object]] = None,
) -> List[Operation]:
    operations: List[Operation] = []
    report_ref = report if report is not None else {}
    report_ref.setdefault("source", "sber")
    report_ref.setdefault("rows_total", 0)
    report_ref.setdefault("skipped", 0)
    report_ref.setdefault("errors", [])

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        headers: Dict[str, int] = {}
        header_row_index = 0
        for row_index, row in enumerate(rows, start=1):
            candidate = _header_index(row)
            if {"Дата", "Тип операции", "Сумма", "Описание"}.issubset(candidate):
                headers = candidate
                header_row_index = row_index
                break

        if not headers:
            _push_skip(report_ref, 1, "header row not found")
            report_ref["imported"] = 0
            return operations

        for row_index, row in enumerate(rows, start=header_row_index + 1):
            if not any(value not in (None, "") for value in row):
                continue
            report_ref["rows_total"] = int(report_ref["rows_total"]) + 1

            op_date = _parse_sber_date(_cell(row, headers, "Дата"))
            if not op_date:
                _push_skip(report_ref, row_index, "invalid date")
                continue

            state_raw = _text(_cell(row, headers, "Состояние"))
            if _is_inactive(state_raw):
                _push_skip(report_ref, row_index, "inactive operation")
                continue

            amount = _parse_amount(_cell(row, headers, "Сумма в рублях")) or _parse_amount(_cell(row, headers, "Сумма"))
            op_type = _detect_type(_text(_cell(row, headers, "Тип операции")), amount)
            signed_amount = amount.copy_abs() if op_type == OperationType.INCOME else -amount.copy_abs()
            account_number = _text(_cell(row, headers, "Номер счета/карты списания")) or "Sber"
            description = _text(_cell(row, headers, "Описание"))
            category = _text(_cell(row, headers, "Категория")) or None
            account_id = vault.ensure_account(bank="sber", name="Сбер", number=account_number)

            operation = Operation(
                id=str(uuid4()),
                account_id=account_id,
                bank="sber",
                date=op_date,
                amount=signed_amount,
                currency=_text(_cell(row, headers, "Валюта")).upper() or "RUB",
                type=op_type,
                description=description,
                merchant=description or None,
                mcc=None,
                bank_category=category,
                source_file_id=file_id,
            )
            if vault.add_operation(operation):
                operations.append(operation)
            else:
                report_ref["duplicates"] = int(report_ref.get("duplicates", 0)) + 1
                _push_skip(report_ref, row_index, "duplicate operation")
    finally:
        workbook.close()

    report_ref["imported"] = len(operations)
    return operations
